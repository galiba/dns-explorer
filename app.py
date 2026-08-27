"""DNS Explorer — dnsdumpster-style DNS recon over a chosen (public or internal) resolver.

Enumeration is done the right way: try a zone transfer (AXFR) against the domain's
nameservers (or a user-supplied internal NS) to dump every record it holds. No
subdomain wordlist — so wildcard DNS can't manufacture fake hosts. If AXFR is
refused, fall back to direct record-type queries on the apex.
"""
import io, ipaddress, secrets, time, json
from concurrent.futures import ThreadPoolExecutor

import dns.resolver, dns.reversename, dns.query, dns.zone, dns.rdatatype
import requests, urllib3
urllib3.disable_warnings()
from fastapi import FastAPI
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = FastAPI(title="DNS Explorer")

APEX_TYPES = ("SOA", "NS", "A", "AAAA", "MX", "TXT", "CNAME", "SRV", "CAA")

_ASN_CACHE: dict[str, str] = {}


class Query(BaseModel):
    domain: str
    resolver: str | None = None   # custom (internal) nameserver IP(s); None = system
    grab_http: bool = True
    passive: bool = True    # crt.sh certificate-transparency discovery


def _split(s: str) -> list[str]:
    return [x.strip() for x in (s or "").replace(",", " ").split() if x.strip()]


def _resolver(ns: str | None) -> dns.resolver.Resolver:
    r = dns.resolver.Resolver(configure=ns is None)
    if ns:
        r.nameservers = _split(ns)
    r.lifetime = r.timeout = 4.0
    return r


def _asn(ip: str) -> str:
    """Team Cymru DNS whois — free, no key. Skips private/IPv6; caches by /24."""
    try:
        if ":" in ip or ipaddress.ip_address(ip).is_private:
            return ""
        key = ip.rsplit(".", 1)[0]
        if key in _ASN_CACHE:
            return _ASN_CACHE[key]
        pub = dns.resolver.Resolver(configure=False)
        pub.nameservers = ["1.1.1.1", "8.8.8.8"]
        pub.lifetime = pub.timeout = 3.0
        rev = ".".join(reversed(ip.split(".")))
        r1 = pub.resolve(f"{rev}.origin.asn.cymru.com", "TXT")[0].to_text().strip('"')
        num, net = r1.split("|")[0].strip(), r1.split("|")[1].strip()
        name = pub.resolve(f"AS{num}.asn.cymru.com", "TXT")[0].to_text().strip('"')
        out = f"{name.split('|')[-1].strip()}\n{num} / {net}"
        _ASN_CACHE[key] = out
        return out
    except Exception:
        return ""


def _ptr(ip: str, ns: str | None) -> str:
    try:
        r = _resolver(ns)
        return str(r.resolve(dns.reversename.from_address(ip), "PTR")[0]).rstrip(".")
    except Exception:
        return ""


def _banner(host: str) -> str:
    out = []
    for scheme in ("http", "https"):
        try:
            r = requests.head(f"{scheme}://{host}", timeout=3, allow_redirects=False,
                              verify=False, headers={"User-Agent": "dns-explorer"})
            out.append(f"{scheme.upper()}: {r.headers.get('Server', 'unknown')} [{r.status_code}]")
        except Exception:
            pass
    return "\n".join(out)


def _enrich(row: dict, ns: str | None, grab_http: bool) -> dict:
    ip = row["value"]
    try:
        row["scope"] = "internal" if ipaddress.ip_address(ip).is_private else "public"
    except Exception:
        row["scope"] = ""
    # reverse-resolve internal IPs via the internal NS, public IPs via system DNS
    ptr_ns = ns if row["scope"] == "internal" else None
    row["reverse"] = _ptr(ip, ptr_ns)
    row["owner"] = _asn(ip)
    row["http"] = _banner(row["host"]) if grab_http else ""
    return row


def _axfr(domain: str, ns_ip: str) -> list[dict]:
    """Zone transfer against one nameserver IP. Returns [] on refusal/error."""
    rows = []
    z = dns.zone.from_xfr(dns.query.xfr(ns_ip, domain, lifetime=8.0))
    for name, node in z.nodes.items():
        fqdn = str(name.derelativize(z.origin)).rstrip(".")
        for rds in node.rdatasets:
            rtype = dns.rdatatype.to_text(rds.rdtype)
            for rd in rds:
                rows.append({"host": fqdn, "type": rtype, "value": rd.to_text(),
                             "reverse": "", "owner": "", "http": "", "source": "AXFR", "view": ""})
    return rows


def _resolve_host(host: str, ns: str | None) -> list[dict]:
    """Resolve one host for A/AAAA/CNAME with its own resolver (thread-safe)."""
    out = []
    r = _resolver(ns)
    for rtype in ("A", "AAAA", "CNAME"):
        try:
            for rec in r.resolve(host, rtype):
                val = rec.address if rtype != "CNAME" else str(rec.target).rstrip(".")
                out.append({"host": host, "type": rtype, "value": val,
                            "reverse": "", "owner": "", "http": ""})
        except Exception:
            pass
    return out


_UA = {"User-Agent": "dns-explorer"}


def _clean(names, domain: str) -> set:
    out = set()
    for n in names:
        n = (n or "").strip().lstrip("*.").lower().rstrip(".")
        if n.endswith(domain) and " " not in n and "@" not in n:
            out.add(n)
    return out


def _src_crtsh(domain: str) -> set:
    # crt.sh frequently 502s / returns HTML under load — retry a couple times
    for attempt in range(3):
        r = requests.get(f"https://crt.sh/?q=%25.{domain}&output=json", timeout=15, headers=_UA)
        if r.ok and r.headers.get("content-type", "").startswith("application/json"):
            out = set()
            for row in r.json():
                out |= _clean(row.get("name_value", "").splitlines(), domain)
            return out
        if attempt < 2:
            time.sleep(1.5)
    return set()


def _src_certspotter(domain: str) -> set:
    r = requests.get(f"https://api.certspotter.com/v1/issuances?domain={domain}"
                     "&include_subdomains=true&expand=dns_names", timeout=15, headers=_UA)
    out = set()
    if r.ok:
        for row in r.json():
            out |= _clean(row.get("dns_names", []), domain)
    return out


def _src_certkit(domain: str) -> set:
    r = requests.get(f"https://ct.certkit.io/search?domain={domain}", timeout=15, headers=_UA)
    out = set()
    if r.ok:
        for row in r.json().get("results", []):
            out |= _clean([row.get("commonName")] + (row.get("dnsNames") or []), domain)
    return out


def _src_hackertarget(domain: str) -> set:
    r = requests.get(f"https://api.hackertarget.com/hostsearch/?q={domain}", timeout=15, headers=_UA)
    if r.ok and "," in r.text and "error" not in r.text.lower():
        return _clean((line.split(",")[0] for line in r.text.splitlines()), domain)
    return set()


_PASSIVE_SOURCES = [_src_crtsh, _src_certspotter, _src_certkit, _src_hackertarget]


def _crtsh(domain: str):
    """Passive subdomain discovery: query ALL certificate-transparency / public
    sources concurrently and union the results — each source finds names the
    others miss. One source failing (e.g. crt.sh 502) never drops the rest.
    Returns (sorted_names, per_source_counts) so flaky sources are visible."""
    def run(fn):
        for attempt in range(2):                 # one retry on hard failure
            try:
                return fn(domain)
            except Exception:
                if attempt == 0:
                    time.sleep(1.0)
        return None                              # None = source errored (vs empty)
    with ThreadPoolExecutor(max_workers=len(_PASSIVE_SOURCES)) as ex:
        results = list(ex.map(run, _PASSIVE_SOURCES))
    names, counts = set(), {}
    for fn, got in zip(_PASSIVE_SOURCES, results):
        label = fn.__name__.replace("_src_", "")
        counts[label] = "err" if got is None else len(got)
        if got:
            names |= got
    return sorted(names), counts


def _run_scan(q: Query):
    """Generator that performs the scan and yields ('stage', msg) progress events
    as each real step happens, then a final ('done', result_dict)."""
    domain = q.domain.strip().rstrip(".").lower()   # normalize: some CT APIs are case-sensitive
    if not domain:
        yield ("done", {"error": "domain required"}); return
    ns = q.resolver
    res = _resolver(ns)
    views = [("internal", ns), ("external", "1.1.1.1 8.8.8.8")] if ns else [("", None)]

    yield ("stage", f"Querying the nameservers for {domain}'s base records (A, AAAA, MX, NS, TXT, SOA)…")
    apex = {rtype: set() for rtype in APEX_TYPES}
    for _, vns in views:
        vr = _resolver(vns)
        for rtype in APEX_TYPES:
            try:
                apex[rtype] |= {r.to_text() for r in vr.resolve(domain, rtype)}
            except Exception:
                pass
    apex = {rtype: sorted(vals) for rtype, vals in apex.items()}

    yield ("stage", "Checking whether the domain uses wildcard DNS…")
    wildcard = False
    try:
        res.resolve(f"{secrets.token_hex(6)}.{domain}", "A")
        wildcard = True
    except Exception:
        pass

    rows, seen = [], {}

    def push(r, source, view):
        k = (r["host"], r["type"], r["value"])
        ex = seen.get(k)
        if ex is not None:
            if view and ex["view"] != view:
                ex["view"] = "both" if ex["view"] else view
            return
        r["source"] = source; r["view"] = view
        seen[k] = r; rows.append(r)

    ns_names = [x.rstrip(".") for x in apex.get("NS", [])]
    targets = []
    for t in _split(ns) or ns_names:
        if any(c.isalpha() for c in t):
            try:
                targets += [(t, str(a)) for a in _resolver(ns).resolve(t, "A")]
            except Exception:
                pass
        else:
            targets.append((t, t))

    yield ("stage", f"Attempting a DNS zone transfer (AXFR) from {len(targets)} nameserver(s)…")
    axfr_from, method = [], ""
    for label, ip in targets:
        try:
            for r in _axfr(domain, ip):
                push(r, "AXFR", r.get("view", ""))
            axfr_from.append(label)
        except Exception:
            continue

    if rows:
        method = "AXFR zone transfer from " + ", ".join(sorted(set(axfr_from)))
        yield ("stage", f"Zone transfer succeeded — parsed {len(rows)} records from the zone.")
    else:
        method = "AXFR refused — direct apex record queries only"
        yield ("stage", "Zone transfer refused — resolving the apex records directly…")
        for rtype in APEX_TYPES:
            if rtype in ("A", "AAAA", "CNAME"):
                continue
            for v in apex.get(rtype, []):
                push({"host": domain, "type": rtype, "value": v,
                      "reverse": "", "owner": "", "http": ""}, "apex", "")
        for label, vns in views:
            for r in _resolve_host(domain, vns):
                push(r, "apex", label)

    passive_names, passive_sources = [], {}
    if q.passive:
        yield ("stage", "Searching certificate transparency logs (crt.sh, Cert Spotter, CertKit, HackerTarget)…")
        all_names, passive_sources = _crtsh(domain)
        cands = [n for n in all_names if n not in {r["host"] for r in rows}][:400]
        passive_names = cands
        yield ("stage", f"Found {len(cands)} candidate hostnames — resolving each to its IP address…")
        jobs = [(h, label, vns) for h in cands for label, vns in views]
        with ThreadPoolExecutor(max_workers=20) as ex:
            results = list(ex.map(lambda j: (j[1], _resolve_host(j[0], j[2])), jobs))
        for label, res_rows in results:
            for r in res_rows:
                push(r, "passive", label)

    a_rows = [r for r in rows if r["type"] in ("A", "AAAA")]
    verb = "Reverse DNS, ASN owner and HTTP banners" if q.grab_http else "Reverse DNS and ASN owner"
    yield ("stage", f"{verb} for {len(a_rows)} IP address(es)…")
    with ThreadPoolExecutor(max_workers=20) as ex:
        list(ex.map(lambda r: _enrich(r, ns, q.grab_http), a_rows))

    yield ("stage", "Flagging internal vs public IPs and assembling the results…")
    if len(views) > 1:
        method += " · combined internal+external resolver"
    if q.passive:
        method += f" + passive CT/DNS ({len(passive_names)} names)"
    rows.sort(key=lambda r: (r["host"], r["type"]))
    counts = {"internal": sum(r.get("scope") == "internal" for r in rows),
              "public": sum(r.get("scope") == "public" for r in rows)}
    yield ("done", {"domain": domain, "resolver": ns or "system", "method": method,
                    "wildcard": wildcard, "counts": counts, "passive_sources": passive_sources,
                    "apex": apex, "rows": rows})


@app.post("/api/scan")
def scan(q: Query):
    def gen():
        try:
            for kind, payload in _run_scan(q):
                key = "stage" if kind == "stage" else "result"
                yield f"data: {json.dumps({key: payload})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'result': {'error': str(e)}})}\n\n"
    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/api/export")
def export(payload: dict):
    domain = payload.get("domain", "domain")
    rows = payload.get("rows", [])
    apex = payload.get("apex", {})
    hdr_font, hdr_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F3B57")

    wb = Workbook()
    s = wb.active; s.title = "Summary"
    s["B2"] = f"DNS Explorer report for {domain}"; s["B2"].font = Font(bold=True, size=14)
    s["B3"] = payload.get("method", ""); s["B3"].font = Font(italic=True, color="666666")
    s["B5"] = "Apex records"; s["B5"].font = Font(bold=True)
    r = 6
    for k, v in apex.items():
        for item in (v or []):
            s.cell(r, 2, k); s.cell(r, 3, item); r += 1
    s.column_dimensions["B"].width = 12; s.column_dimensions["C"].width = 70

    d = wb.create_sheet("DNS Records")
    cols = ["Host", "Type", "Value / IP", "Scope", "View", "Reverse DNS", "Netblock Owner", "HTTP Services", "Source"]
    for c, name in enumerate(cols, 1):
        cell = d.cell(1, c, name); cell.font = hdr_font; cell.fill = hdr_fill
    for i, row in enumerate(rows, 2):
        for c, key in enumerate(("host", "type", "value", "scope", "view", "reverse", "owner", "http", "source"), 1):
            d.cell(i, c, row.get(key))
    for col, w in zip("ABCDEFGHI", (34, 8, 40, 10, 9, 30, 44, 40, 10)):
        d.column_dimensions[col].width = w
    wrap = Alignment(wrap_text=True, vertical="top")
    for r_ in d.iter_rows(min_row=2):
        for c in r_:
            c.alignment = wrap

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{domain}-dns.xlsx"'})


app.mount("/", StaticFiles(directory="static", html=True), name="static")
